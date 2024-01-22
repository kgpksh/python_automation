from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'MyNewSheet'

wb.create_sheet('SecondSheet')

wb.create_sheet('ThirdSheet', 1)

ws['A1'] = 1
ws['B1'] = 2
ws['A2'] = 3
ws.cell(row = 2, column = 2, value = 4)

wb.save('./excel_playground/practice2.xlsx')
wb.close()