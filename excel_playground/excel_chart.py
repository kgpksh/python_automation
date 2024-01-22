from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from random import randint

path = './excel_playground/practice4.xlsx'
wb = Workbook()

ws = wb.active
ws.append(['번호', '영어', '수학'])

for i in range(1, 11) :
    ws.append([i, randint(0, 100), randint(0, 100)])

wb.save(path)
wb.close()

from openpyxl.chart import LineChart, BarChart, Reference

wb = load_workbook(path)
ws = wb.active

# B2:C11까지의 데이터를 차트로 생성
chart_value = Reference(ws, min_row = 2, max_row = 11, min_col = 2, max_col = 3)

bar_chart =  BarChart()
bar_chart.add_data(chart_value)
ws.add_chart(bar_chart, 'E1')

line_chart = LineChart()
line_chart.add_data(chart_value)
ws.add_chart(line_chart, 'L1')

# min_row를 1로 바꿔 데이터 범례 나타내기
chart_value = Reference(ws, min_row = 1, max_row = 11, min_col = 2, max_col = 3)
line_chart = LineChart()
line_chart.add_data(chart_value, titles_from_data=True)
line_chart.title = '성적표' # 제목
line_chart.style = 20 # 미리 정의된 스타일 적용. 사용자가 개별 지정도 가능
line_chart.y_axis.title = '점수' # Y축의 제목
line_chart.x_axis.title = '번호' # X 축의 제목
ws.add_chart(line_chart, 'L15')

wb.save(path)