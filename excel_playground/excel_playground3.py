from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string

path = './excel_playground/practice3.xlsx'
wb = Workbook()

ws = wb.active
ws['A1'] = 1
ws['B1'] = 2
ws['A2'] = 3
ws.cell(row = 2, column = 2, value = 4)

wb.save(path)
wb.close()


# 값 순회하면서 읽어오기
wb = load_workbook(path)
ws = wb.active
for row in ws :
    for column in row :
        print(column.value, end = ' ')
    print()

# cell 함수로 값 읽어오기
print('----------------------')
for x in range(1, ws.max_row + 1) :
    for y in range(1, ws.max_column + 1) :
        print(ws.cell(row = x, column = y).value, end = ' ')
    print()

# 범위로 값 읽어오기
print('----------------------')
for row in ws[1:2]:
    for col in row :
        print(col.value, end = ' ')
    print()

# 좌표 가져오기
print('----------------------')
for row in ws[1:2]:
    for col in row :
        print(col.coordinate, end = ' ')
    print()

# 튜플로 좌표 가져오기
print('----------------------')
for row in ws[1:2]:
    for col in row :
        xy = coordinate_from_string(col.coordinate)
        print(xy, end = ' ')
    print()

print('----------------------')
# row들을 기준으로 묶어 불러오기
for row in tuple(ws.rows) :
    print(row)

print('----------------------')
# column들을 기준으로 묶어 불러오기
for col in tuple(ws.columns) :
    print(col)

print('----------------------')
# 슬라이싱
for row in ws.iter_rows(min_row = 1, max_row = 4, min_col = 2, max_col = 6) :
    print(row[0].value)


# 2번열을 기준으로 3칸 밀기
ws.insert_cols(2, 3)
# 1번행을 기준으로 2칸 밀기
ws.insert_rows(1, 2)
# 4번행을 기준으로 1칸 추가
ws.insert_rows(4)

# 1번행을 기준으로 2칸 제거
ws.delete_rows(1, 2)

# A1 부터 B2까지 드래그하여 행 1, 열 2칸 옮기기
ws.move_range('A1:B2', rows = 1, cols = 2)
# B:2자리에 값 추가
ws['B2'].value = 'new Value'

# move같은 기능을 사용 시 기존에 있던 데이터를 덮어쓰게 될 수 있다.
wb.save(path)